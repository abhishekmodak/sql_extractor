from openpyxl import load_workbook, Workbook
import datetime

st = datetime.datetime.now()
SQL_FUNCTIONS = [
                    'AVG(', 'COUNT(', 'FIRST(', 'LAST(', 'MAX(', 'MIN(', 'SUM(',
                    'UCASE(', 'LCASE(', 'MID(', 'LEN(', 'ROUND(', 'NOW(', 'FORMAT('
                     'avg(', 'count(', 'first(', 'last(', 'max(', 'min(', 'sum(',
                    'ucase(', 'lcase(', 'mid(', 'len(', 'round(', 'now(', 'format('
            ]
function_list = []
normal_list = []
wb = load_workbook(filename='DB.xlsx', read_only=True)
ws = wb['Sheet1']
wb1 = Workbook(write_only=True)
wb2 = Workbook(write_only=True)
ws1 = wb1.create_sheet()
ws2 = wb2.create_sheet()

for row in ws.rows:
    for cell in row:
        if cell.value and ("select" in cell.value or "Select" in cell.value):
            cell_flag = True
            for func in SQL_FUNCTIONS:
                if func in cell.value:
                    function_list.append(cell.value)
                    cell_flag = False
                    continue
            if cell_flag:
                normal_list.append(cell.value)

        #else:
            #print("No")

#print("########## Function List ######")
#print(function_list)
#print("********* Normal List ********")
#print(normal_list)
et = datetime.datetime.now()

print("222222222222")
print(et-st)
print('writing functions')
for item in function_list:
    ws1.append([item])
et = datetime.datetime.now()

print("1111111111111")
print(et-st)

print('writing normal')
for item in normal_list:
    ws2.append([item])
wb1.save('function.xlsx')
wb2.save('normal.xlsx')
et = datetime.datetime.now()

print("xxxxxxxxxxx")
print(et-st)
